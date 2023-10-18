import babel
import openpyxl
from django.shortcuts import get_object_or_404
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from .models import *
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
from .models import BuktiKasKeluar
from babel.numbers import format_currency
from decimal import Decimal
from django.contrib import admin
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.units import mm, cm
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
from reportlab.lib.styles import getSampleStyleSheet
import babel.numbers
import decimal
from django.contrib.admin.widgets import FilteredSelectMultiple
import os
from django.conf import settings
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import cm
from babel.numbers import format_currency
from reportlab.lib.units import inch
from terbilang import Terbilang
from rangefilter.filters import DateRangeFilterBuilder
from openpyxl.styles import Font



styles = getSampleStyleSheet()

class DanaMasukAdmin(admin.ModelAdmin):
    search_fields = ('waktu_masuk','uraian', 'bank_penerima__nomer_bank_tertarik',)
    list_display = ('waktu_masuk', 'uraian', 'bank_penerima', 'get_total_dana')
    actions = ["export_as_pdf", "export_to_excel","update_50_list"]
    list_per_page = 20
    list_filter = (("waktu_masuk", DateRangeFilterBuilder()),)


    def changelist_view(self, request, extra_context=None):
        if 'per_page' in request.GET:
            per_page = int(request.GET['per_page'])
            if per_page > 0:
                self.list_per_page = per_page
            else:
                self.list_per_page = self.list_max_show_all
        return super().changelist_view(request, extra_context=extra_context)

    def export_to_excel(self, request, queryset):
        # Mendapatkan kolom yang akan diekspor
        data = queryset.values('waktu_masuk', 'uraian', 'bank_penerima', 'total_dana')

        # Membuat file Excel
        wb = Workbook()
        ws = wb.active
        ws.append(['Waktu Masuk', 'Uraian', 'Bank Penerima', 'Total Dana'])

        row_num = 1

        # Menambahkan header dengan nama kolom
        for item in data:
            row = [row_num, item['waktu_masuk'], item['uraian'], item['bank_penerima'], item['total_dana'],]
            ws.append(row)
            row_num += 1


        # Mengatur respons HTTP
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=DanaMasuk.xlsx'

        # Simpan file Excel ke respons
        wb.save(response)

        return response

    export_to_excel.short_description = 'Export to Excel'


    def export_as_pdf(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Dana Masuk.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = TA_CENTER
        title = Paragraph('Dana Masuk', style=title_style)
        elements.append(title)
        data = []
        data.append(['No.','Uraian', 'Waktu Masuk', 'Total Dana'])
        total = Decimal(0)
        row_num = 1
        for dana_masuk in queryset:
            total_ajuan = Decimal(dana_masuk.total_dana)
            row = [
                row_num,
                dana_masuk.uraian,
                dana_masuk.waktu_masuk,
                format_currency(dana_masuk.total_dana, 'IDR', locale='id_ID'),
            ]
            data.append(row)
            row_num += 1

            # Add total_ajuan to total
            total += total_ajuan

            # Add row for total_ajuan
        data.append(['Total', '', '','', format_currency(total, 'IDR', locale='id_ID')])

        table = Table(data, colWidths=[70, 200, 150, 150,])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -2), 'LEFT'),  # Set "NAMA KEGIATAN" in the second row to align left
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('WORDWRAP', (1, 0), (-1, -1), 100),
        ]))

        elements.append(table)
        doc.build(elements)

        return response

    export_as_pdf.short_description = "Export selected as PDF"

    def get_total_dana(self, obj):
        return babel.numbers.format_currency(obj.total_dana, 'IDR', locale='id_ID')
    get_total_dana.short_description = 'Total Dana'



class BuktiKasKeluarAdmin(admin.ModelAdmin):
    readonly_fields = ['no_BKK', ]
    fields = ['no_BKK', 'tanggal_BKK', 'ajuan', 'dibayarkan_kepada', 'uraian', 'nomer_cek','nomer_bank_tertarik']
    search_fields = ['no_BKK','ajuan__nomor_pengajuan', 'tanggal_BKK', 'dibayarkan_kepada', 'uraian','nomer_bank_tertarik__nomer_bank_tertarik','nomer_cek__no_cek']
    list_display = ('no_BKK', 'tanggal_BKK', 'ajuan','get_total_ajuan', 'dibayarkan_kepada', 'uraian', 'nomer_bank_tertarik', 'nomer_cek')
    actions = ["export_to_excel", "export_as_pdf_global", "export_to_pdf_satuan","test_add_logo"]
    raw_id_fields = ['ajuan', 'nomer_cek']
    list_per_page = 20
    list_filter = (("tanggal_BKK", DateRangeFilterBuilder()),)

    def changelist_view(self, request, extra_context=None):
        if 'per_page' in request.GET:
            per_page = int(request.GET['per_page'])
            if per_page > 0:
                self.list_per_page = per_page
            else:
                self.list_per_page = self.list_max_show_all
        return super().changelist_view(request, extra_context=extra_context)


    def set_table_style(self, table):
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),  # Mengubah warna latar belakang header menjadi putih
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('TOPPADDING', (0, 0), (-1, 0), 2.5),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Mengubah warna latar belakang sel lainnya menjadi putih
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

    def generate_first_table(self, elements, data):
        def wrap_text_by_words(text, limit=7):
            words = text.split()
            return '\n'.join([' '.join(words[i:i + limit]) for i in range(0, len(words), limit)])

        logo_path = os.path.join(settings.STATIC_ROOT, 'mysite/logo rahmany.png')
        logo = Image(logo_path)
        logo.drawHeight = 1.5 * cm
        logo.drawWidth = 2 * cm

        total_ajuan = data.ajuan.total_ajuan if data and data.ajuan and data.ajuan.total_ajuan else 0
        total_ajuan_str = str(total_ajuan)
        t = Terbilang()
        t.parse(total_ajuan_str)
        t_gr = t.getresult()
        t_gr_string_title = t_gr.title() + ' Rupiah'
        t_gr_string_title = wrap_text_by_words(t_gr_string_title, 7)


        data_0 = [
            [logo, '\n BUKTI KAS KELUAR\n_________________',
             'Nomer BKK: {}\nTanggal: {}'.format(data.no_BKK, data.tanggal_BKK) if data else ''],
            ['Perkiraan', 'Uraian', 'Jumlah'],
            ['', '', ''],
            ['', '', ''],
            ['Dibayarkan Kepada:\n {}'.format(data.dibayarkan_kepada), data.uraian,
             format_currency(data.ajuan.total_ajuan, 'IDR', locale='id_ID') if data.ajuan.total_ajuan else ''],
            ['Terbilang', '{}'.format(t_gr_string_title), '']
        ]

        table_0 = Table(data_0, colWidths=[150, 250, 150])
        self.set_table_style(table_0)
        elements.append(table_0)


    def generate_second_table(self, elements, data):
        data_2 = [
            ['Nomer Bank: {}'.format(
                data.nomer_bank_tertarik.nomer_bank_tertarik if data.nomer_bank_tertarik else ''),
             'Nomer Cek: {}'.format(data.nomer_cek.no_cek if data.nomer_cek else '')]
        ]
        table_2 = Table(data_2, colWidths=[275, 275])
        self.set_table_style(table_2)
        elements.append(table_2)

    def generate_signature_table(self, elements, data):
        tanda_tangan = [['Pemberi \n \n \n____________', 'Mengetahui \n \n \n____________',
                         'Penerima \n \n \n____________']]
        table_ttd = Table(tanda_tangan, colWidths=[175, 200, 175])
        self.set_table_style(table_ttd)
        elements.append(table_ttd)



    def export_to_pdf_satuan(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="BKK Satuan.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.1*inch)


        elements = []
        for data in queryset:
            # Kode untuk tabel pertama
            self.generate_first_table(elements, data)

            # Kode untuk tabel kedua
            self.generate_second_table(elements, data)

            # Kode untuk tanda tangan
            self.generate_signature_table(elements, data)

            # Kode untuk spacer
            data_spacer = [('')]
            table_spacer = Table(data_spacer, colWidths=[500])
            elements.append(table_spacer)

        doc.build(elements)

        return response

    export_to_pdf_satuan.short_description = "Export selected as PDF satuan"

    def get_total_ajuan(self, obj):
        if hasattr(obj, 'ajuan') and hasattr(obj.ajuan, 'total_ajuan'):
            total_ajuan = decimal.Decimal(str(obj.ajuan.total_ajuan))
        else:
            total_ajuan = decimal.Decimal('0')
        return babel.numbers.format_currency(total_ajuan, 'IDR', locale='id_ID')
    get_total_ajuan.short_description = 'Total Ajuan'

    def export_to_excel(self, request, queryset):
        # Query data dari model BuktiKasKeluar
        data = queryset.values('no_BKK', 'tanggal_BKK', 'ajuan_id', 'dibayarkan_kepada', 'uraian',
                               'nomer_bank_tertarik', 'nomer_cek')

        # Buat file Excel dan tambahkan header
        wb = Workbook()
        ws = wb.active
        ws.append(['No.', 'No BKK', 'Tanggal BKK', 'Ajuan', 'Dibayarkan Kepada', 'Uraian', 'Nomer Bank',
                   'Nomor Cek'])
        row_num = 1

        # Tambahkan data ke file Excel
        for item in data:
            row = [row_num, item['no_BKK'], item['tanggal_BKK'], item['ajuan_id'], item['dibayarkan_kepada'],
                   item['uraian'], item['nomer_bank_tertarik'], item['nomer_cek']]
            ws.append(row)
            row_num += 1

        # Konversi file Excel ke HttpResponse
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=BuktiKasKeluar.xlsx'
        wb.save(response)

        return response

    export_to_excel.short_description = 'Export to Excel'

    def export_as_pdf_global(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="BKK Global.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        title_style = styles['Heading1']
        title_style.alignment = TA_CENTER
        title = Paragraph('Bukti Kas Keluar', style=title_style)
        elements.append(title)
        data = []
        data.append(
            ['No.', 'Nomor BKK', 'Tanggal BKK', 'Nomor Cek', 'Dibayarkan Kepada', 'Uraian', 'Kode Bank', 'Ajuan'])
        total = Decimal(0)
        row_num = 1
        for bukti_kas_keluar in queryset:
            total_ajuan = Decimal(bukti_kas_keluar.ajuan.total_ajuan) if bukti_kas_keluar.ajuan else Decimal(0)
            row = [
                row_num,
                bukti_kas_keluar.no_BKK,
                bukti_kas_keluar.tanggal_BKK,
                bukti_kas_keluar.dibayarkan_kepada,
                bukti_kas_keluar.uraian,
                bukti_kas_keluar.nomer_bank_tertarik,
                bukti_kas_keluar.nomer_cek,
                format_currency(total_ajuan, 'IDR', locale='id_ID') if total_ajuan else '-',
            ]
            data.append(row)
            row_num += 1

            # Add total_ajuan to total
            total += total_ajuan

        # Add row for total_ajuan
        data.append(['', '', '', '', '', '', 'Total', format_currency(total, 'IDR', locale='id_ID')])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 2), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (0, 0), (7, 0), 'CENTER'),
            ('ALIGN', (7, 1), (7, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('WORDWRAP', (1, 0), (-1, -1), 100),
        ]))

        elements.append(table)
        doc.build(elements)

        return response

    export_as_pdf_global.short_description = "Export selected as Global PDF"



class AjuanAdmin(admin.ModelAdmin):
    search_fields = ('nomor_pengajuan', 'nama_kegiatan', 'waktu_ajuan', 'total_ajuan',)
    list_display = ['unit_ajuan','nomor_pengajuan',  'nama_kegiatan', 'waktu_ajuan', 'get_total_ajuan', 'penanggung_jawab', 'RAPT',  ]
    fields = ['unit_ajuan','nomor_pengajuan',  'nama_kegiatan', 'waktu_ajuan', 'total_ajuan', 'penanggung_jawab', 'RAPT',  ]
    readonly_fields = ['nomor_pengajuan', ]
    raw_id_fields = ['RAPT',]
    actions = ['export_as_pdf','export_to_excel']
    list_per_page = 20  # Jumlah item per halaman default
    list_filter = (("waktu_ajuan", DateRangeFilterBuilder()),)

    def changelist_view(self, request, extra_context=None):
        if 'per_page' in request.GET:
            per_page = int(request.GET['per_page'])
            if per_page > 0:
                self.list_per_page = per_page
            else:
                self.list_per_page = self.list_max_show_all
        return super().changelist_view(request, extra_context=extra_context)

    def get_total_ajuan(self, obj):
        return babel.numbers.format_currency(obj.total_ajuan, 'IDR', locale='id_ID')
    get_total_ajuan.short_description = 'Total Ajuan'

    def export_to_excel(self, request, queryset):
        # Query data dari model BuktiKasKeluar
        data = queryset.values('unit_ajuan__unit_ajuan', 'nomor_pengajuan', 'nama_kegiatan', 'waktu_ajuan', 'total_ajuan', 'RAPT')

        # Buat file Excel dan tambahkan header
        wb = Workbook()
        ws = wb.active
        ws.append(['No.','Unit Ajuan','Nomor Pengajuan','Nama Kegiatan','Waktu Ajuan','Total Ajuan','RAPT'])
        # set rata tengah pada header
        header_row = ws[1]
        for cell in header_row:
            cell.alignment = Alignment(horizontal='center')

        row_num = 1
        # Tambahkan data ke file Excel
        for item in data:
            row = [row_num, item['unit_ajuan__unit_ajuan'], item['nomor_pengajuan'], item['nama_kegiatan'], item['waktu_ajuan'], item['total_ajuan'],
                   item['RAPT']]
            ws.append(row)
            row_num += 1

            # Menyesuaikan lebar kolom secara otomatis
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length

        # Konversi file Excel ke HttpResponse
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=Ajuan.xlsx'
        wb.save(response)

        return response

    export_to_excel.short_description = 'Export to Excel'

    def export_as_pdf(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="ajuan.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        title_style = styles['Heading1']
        title_style.alignment = TA_CENTER
        title = Paragraph('Daftar Ajuan', style=title_style)
        elements.append(title)
        data = []
        data.append(['NO','UNIT AJUAN', 'NOMOR PENGAJUAN', 'NAMA KEGIATAN', 'WAKTU AJUAN', 'TOTAL AJUAN', 'RAPT',])
        total = Decimal(0)
        row_num = 1
        for ajuan in queryset:
            total_ajuan = Decimal(ajuan.total_ajuan)
            row = [
                row_num,
                ajuan.unit_ajuan,
                ajuan.nomor_pengajuan,
                ajuan.nama_kegiatan,
                ajuan.waktu_ajuan,
                format_currency(ajuan.total_ajuan, 'IDR', locale='id_ID'),
                ajuan.RAPT,
            ]
            data.append(row)
            row_num += 1

            # Add total_ajuan to total
            total += total_ajuan

            # Add row for total_ajuan
        data.append(['', '', '', '','Total', format_currency(total, 'IDR', locale='id_ID', )])


        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -2), 'LEFT'),
            ('ALIGN', (2, 1), (2, -2), 'LEFT'),
            ('ALIGN', (1, 1), (1, -2), 'LEFT'),
            ('ALIGN', (5, 1), (5, -2), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('WORDWRAP', (1, 0), (-1, -1), 100),
        ]))

        elements.append(table)
        doc.build(elements)

        return response

    export_as_pdf.short_description = "Export selected as PDF"

class AjuanInLineRAPT(admin.TabularInline):
    model = Ajuan
    extra = 0
    fields = ['unit_ajuan','nomor_pengajuan',  'nama_kegiatan', 'waktu_ajuan', 'total_ajuan', 'penanggung_jawab']
    readonly_fields = fields
    can_delete = False

# class


class RekapAjuanPengambilanTabunganAdmin(admin.ModelAdmin):
    search_fields = ('no_RAPT', 'jumlah')
    list_display = ('no_RAPT','get_nomor_pengajuan','get_total_ajuan', 'get_jumlah')
    readonly_fields = ('no_RAPT','jumlah',)
    inlines = [AjuanInLineRAPT]
    list_per_page = 20  # Jumlah item per halaman default

    def changelist_view(self, request, extra_context=None):
        if 'per_page' in request.GET:
            per_page = int(request.GET['per_page'])
            if per_page > 0:
                self.list_per_page = per_page
            else:
                self.list_per_page = self.list_max_show_all
        return super().changelist_view(request, extra_context=extra_context)

    def get_nomor_pengajuan(self, obj):
        nomor_pengajuan_list = []
        for ajuan in obj.ajuan_set.all():
            nomor_pengajuan_list.append(ajuan.nomor_pengajuan)
        if nomor_pengajuan_list:
            return ", ".join(nomor_pengajuan_list)
        else:
            return '-'

    get_nomor_pengajuan.short_description = 'nomor pengajuan'

    def get_total_ajuan(self, obj):
        total_ajuan_list = []
        for ajuan in obj.ajuan_set.all():
            total_ajuan_list.append(str(ajuan.total_ajuan))  # ubah objek Decimal menjadi str
        if total_ajuan_list:
            return ", ".join(total_ajuan_list)
        else:
            return '-'

    get_total_ajuan.short_description = 'total ajuan'
    actions = ["export_as_pdf", "export_to_excel"]

    def get_jumlah(self, obj):
        return babel.numbers.format_currency(obj.jumlah, 'IDR', locale='id_ID')
    get_jumlah.short_description = 'Jumlah'


    def export_as_pdf(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="RAPT.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(letter))

        data = []
        data.append(['no_RAPT', 'jumlah'])
        total = Decimal(0)
        for RAPT in queryset:
            row = [
                RAPT.no_RAPT,
                RAPT.jumlah
            ]
            data.append(row)

            # Add total_ajuan to total
            total += RAPT.jumlah

            # Add row for total_ajuan
        data.append(['Total', format_currency(total, 'IDR', locale='id_ID')])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))


        elements = []
        elements.append(table)
        doc.build(elements)

        return response

    export_as_pdf.short_description = "Export selected as PDF"

from django.http import HttpResponse
from django.db.models import Q
class CekAdmin(admin.ModelAdmin):
    raw_id_fields = ('RPC',)
    search_fields = ('no_cek','nomer_bank_tertarik__nomer_bank_tertarik','RPC__no_RPC',)
    list_display = ('tanggal', 'no_cek', 'nomer_bank_tertarik', 'display_many_to_many', 'get_total_cek', 'RPC',)
    actions = ["export_as_pdf","export_as_excel","save_all_selected"]
    formfield_overrides = {
        models.ManyToManyField: {'widget': FilteredSelectMultiple('Ajuan', False)},
    }
    list_per_page = 20  # Jumlah item per halaman default
    list_filter = (("tanggal", DateRangeFilterBuilder()),)

    def save_all_selected(modeladmin, request, queryset):
        for obj in queryset:
            obj.save()

    save_all_selected.short_description = "Update Data"

    def changelist_view(self, request, extra_context=None):
        if 'per_page' in request.GET:
            per_page = int(request.GET['per_page'])
            if per_page > 0:
                self.list_per_page = per_page
            else:
                self.list_per_page = self.list_max_show_all
        return super().changelist_view(request, extra_context=extra_context)

    def queryset(self, request):
        # Menggunakan method queryset() untuk menyesuaikan query yang digunakan dalam list_display
        qs = super().queryset(request)
        # Filter hanya objek-objek yang memiliki total_cek = 0
        return qs.filter(total_cek=0)

    def get_total_cek(self, obj):
        return babel.numbers.format_currency(obj.total_cek, 'IDR', locale='id_ID')
    def get_form(self, request, obj=None, **kwargs):
        request._obj_ = obj
        return super(CekAdmin, self).get_form(request, obj, **kwargs)
    def formfield_for_manytomany(self, db_field, request, **kwargs):
        if db_field.name == "ajuan_terkait":
            if hasattr(request, '_obj_') and request._obj_ is not None:
                kwargs['queryset'] = Ajuan.objects.filter(
                    Q(is_selected=False) | Q(ceks_terkait=request._obj_)
                )
            else:

                kwargs['queryset'] = Ajuan.objects.filter(is_selected=False)
        return super().formfield_for_manytomany(db_field, request, **kwargs)



    def display_many_to_many(self, obj):
        # Ambil nomor_pengajuan dari setiap objek Ajuan yang terkait dalam ajuan_terkait
        nomor_pengajuan_list = [str(ajuan.nomor_pengajuan) for ajuan in obj.ajuan_terkait.all()]

        # Menggabungkan nomor_pengajuan menjadi satu string dengan koma dan spasi sebagai pemisah
        return ', '.join(nomor_pengajuan_list)
    display_many_to_many.short_description = 'Ajuan Terkait'

    def export_as_pdf(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Cek.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []

        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = 1  # TA_CENTER is 1
        title = Paragraph('Daftar Cek', style=title_style)
        elements.append(title)

        row_num = 1
        data = [['No.', 'Tanggal', 'No.Cek', 'Nomer Bank','Ajuan', 'RPC', 'Total Ajuan']]

        total_sum = 0  # Inisialisasi total

        for cek in queryset:
            if cek.total_cek is not None:
                formatted_total_cek = format_currency(cek.total_cek, 'IDR', locale='id_ID')
            else:
                formatted_total_cek = "-"  # Or any string you want to display when total_cek is None
            ajuans = ", ".join(str(ajuan.nomor_pengajuan) for ajuan in cek.ajuan_terkait.all())
            total_ajuan = cek.ajuan_terkait.aggregate(Sum('total_ajuan'))[
                'total_ajuan__sum']  # asumsikan total_ajuan adalah field dalam model Ajuan

            # Menambahkan 'cek.tanggal' dan 'total_ajuan' ke baris tabel
            row = [
                row_num,
                cek.tanggal,
                cek.no_cek,
                cek.nomer_bank_tertarik,
                ajuans,
                cek.RPC,
                formatted_total_cek,  # Total Ajuan

            ]
            data.append(row)
            row_num += 1

            if cek.total_cek is not None:
                total_sum += cek.total_cek

        total_row = [
            'Total', '', '', '', '', '', format_currency(total_sum, 'IDR', locale='id_ID')
        ]
        data.append(total_row)

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
            ('ALIGN', (0, 1), (1, -2), 'CENTER'),
            ('ALIGN', (0, 0), (6, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('WORDWRAP', (1, 0), (-1, -1), 100),
        ]))
        elements.append(table)
        doc.build(elements)
        return response

    export_as_pdf.short_description = "Export selected as PDF"

    def export_as_excel(self, request, queryset):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Cek.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cek Data'

        columns = ['No.', 'Tanggal', 'No.Cek', 'Nomer Bank', 'Ajuan', 'RPC', 'Total Ajuan']

        font_bold = Font(bold=True)

        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = font_bold

        row_num = 2
        total_sum = 0

        for cek in queryset:
            if cek:
                ajuans = ", ".join(str(ajuan.nomor_pengajuan) for ajuan in cek.ajuan_terkait.all())
                if cek.total_cek is not None:
                    formatted_total_cek = format_currency(cek.total_cek, 'IDR', locale='id_ID')
                    total_sum += cek.total_cek
                else:
                    formatted_total_cek = "-"

                # Check if 'nomer_bank_tertarik' has the 'no_RPC' attribute
                if hasattr(cek.nomer_bank_tertarik, 'no_RPC'):
                    no_rpc = cek.nomer_bank_tertarik.no_RPC
                else:
                    no_rpc = None

                row = [row_num - 1, cek.tanggal, cek.no_cek, no_rpc, ajuans, formatted_total_cek]

                # Iterate over the row data and write it to the worksheet
                for col_num, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = cell_value

                row_num += 1

        total_row = ['Total', '', '', '', '', '', format_currency(total_sum, 'IDR', locale='id_ID')]
        for col_num, cell_value in enumerate(total_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.font = font_bold

        response['Content-Disposition'] = 'attachment; filename="Cek.xlsx"'
        wb.save(response)

        return response

    export_as_excel.short_description = "Export selected as Excel"


class CekInLineRPC(admin.TabularInline):
    model = Cek
    extra = 0
    fields = ['no_cek', 'nomer_bank_tertarik','get_nama_kegiatan',]
    readonly_fields = fields
    can_delete = False

    def get_nama_kegiatan(self, obj):
        return obj.ajuan.nama_kegiatan

    get_total_nama_kegiatan_description = 'Nama Kegiatan'

from reportlab.pdfgen import canvas

class RekapPencairanCekAdmin(admin.ModelAdmin):
    search_fields = ('no_RPC','jumlah')
    readonly_fields = ('no_RPC', 'jumlah',)
    inlines = [CekInLineRPC]
    actions = ['export_as_pdf','export_as_excel', 'update_jumlah_RPC', 'export_as_pdf2']
    list_display = ('no_RPC','tanggal_syc_cek', 'get_jumlah','nomor_cek', 'nomer_bank_tertarik', 'get_nama_kegiatan_from_rpc','get_total_ajuan',)
    list_per_page = 20  # Jumlah item per halaman default


    def get_search_results(self, request, queryset, search_term):
        # Pemrosesan pencarian dalam model saat ini (RekapPencairanCek)
        results, use_distinct = super().get_search_results(request, queryset, search_term)

        # Pemrosesan pencarian dalam model Cek berdasarkan ForeignKey (RPC)
        cek_results = Cek.objects.filter(
            Q(nomer_bank_tertarik__nomer_bank_tertarik__icontains=search_term) |
            Q(ajuan_terkait__nama_kegiatan__icontains=search_term)  |
            Q(ajuan_terkait__total_ajuan__icontains=search_term)
        )
        cek_results = cek_results.values_list('RPC__id', flat=True)  # Ambil ID yang sesuai

        # Gabungkan hasil pencarian dari model saat ini dan model Cek
        results |= queryset.filter(id__in=cek_results)

        return results, use_distinct

    def changelist_view(self, request, extra_context=None):
        if 'per_page' in request.GET:
            per_page = int(request.GET['per_page'])
            if per_page > 0:
                self.list_per_page = per_page
            else:
                self.list_per_page = self.list_max_show_all
        return super().changelist_view(request, extra_context=extra_context)

    def get_jumlah(self, obj):
        return babel.numbers.format_currency(obj.jumlah, 'IDR', locale='id_ID')
    get_jumlah.short_description = 'Jumlah'

    def update_jumlah_RPC(self, request, queryset):
        for obj in queryset:
            total_ajuan = 0
            for cek in Cek.objects.filter(RPC=obj):
                total_ajuan += sum(ajuan.total_ajuan for ajuan in cek.ajuan_terkait.all())
            obj.jumlah = total_ajuan
            obj.save(update_fields=['jumlah'])
        self.message_user(request, f'Successfully updated {queryset.count()} Rekap Pencairan Cek(s).')

    update_jumlah_RPC.short_description = 'Update Jumlah RPC'

    def tanggal_syc_cek(self, rpc_obj):
        ceks_related_to_rpc = Cek.objects.filter(RPC=rpc_obj)
        if ceks_related_to_rpc.exists():
            return ceks_related_to_rpc.first().tanggal  # Mengambil tanggal dari objek pertama dalam queryset
        else:
            return None  # Mengembalikan None jika tidak ada objek Cek yang terkait
    tanggal_syc_cek.short_description = 'Tanggal'

    def get_nama_kegiatan_from_rpc(self, rpc_obj):
        nama_kegiatan_list = []

        # Akses semua objek Cek yang terkait dengan objek RPC
        ceks_related_to_rpc = Cek.objects.filter(RPC=rpc_obj)

        # Loop melalui semua objek Cek dan akses ajuannya
        for cek in ceks_related_to_rpc:
            for ajuan in cek.ajuan_terkait.all():
                nama_kegiatan_list.append(ajuan.nama_kegiatan)

        if nama_kegiatan_list:
            return ", ".join(nama_kegiatan_list)
        else:
            return '-'

    get_nama_kegiatan_from_rpc.short_description = 'Nama Kegiatan'

    def get_total_ajuan(self, rpc_obj):
        ceks_related_to_rpc = Cek.objects.filter(RPC=rpc_obj)
        total_ajuan_list = []

        for cek in ceks_related_to_rpc:
            for ajuan in cek.ajuan_terkait.all():
                total_ajuan_list.append(str(ajuan.total_ajuan))

        if total_ajuan_list:
            return ", ".join(total_ajuan_list)
        else:
            return '-'

    get_total_ajuan.short_description = 'Total Ajuan'


    def get_nama_kegiatan(self, obj):
        nama_kegiatan_list = []

        for ajuan in obj.ceks_terkait.all():
            nama_kegiatan_list.append(ajuan.nama_kegiatan)

        if nama_kegiatan_list:
            return ", ".join(nama_kegiatan_list)
        else:
            return '-'

    get_nama_kegiatan.short_description = 'Nama Kegiatan'


    def nomer_bank_tertarik(self, obj):
        nomer_bank_tertarik_list = []
        for cek in obj.cek_set.all():
            nomer_bank_tertarik_list.append(str(cek.nomer_bank_tertarik))
        if nomer_bank_tertarik_list:
            return ", ".join(nomer_bank_tertarik_list)
        else:
            return '-'
    nomer_bank_tertarik.short_description = 'Bank'

    def nomor_cek(self, obj):
        nomor_cek = []
        for cek in obj.cek_set.all():
            nomor_cek.append(str(cek.no_cek))
        if nomor_cek:
            return ", ".join(nomor_cek)
        else:
            return '-'
    nomer_bank_tertarik.short_description = 'No Cek'


    def save_model(self, request, obj, form, change):
        if obj.pk:
            total_ajuan = 0
            # Dapatkan semua objek Cek yang memiliki foreign key ke objek RekapPencairanCek ini
            for cek in Cek.objects.filter(RPC=obj):
                # Jumlahkan semua total_ajuan dari objek Ajuan yang terkait dengan objek Cek ini
                total_ajuan += sum(ajuan.total_ajuan for ajuan in cek.ajuan_terkait.all())
            # Tetapkan total_ajuan sebagai jumlah di objek RekapPencairanCek ini
            obj.jumlah = total_ajuan
        super().save_model(request, obj, form, change)

    def export_as_pdf(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="RPC.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=10, rightMargin=5, topMargin=5, bottomMargin=5)
        elements = []

        # Add title
        title_style = getSampleStyleSheet()['Heading1']
        title_style.alignment = TA_CENTER
        title = Paragraph('Daftar RPC' , style=title_style)
        elements.append(title)

        for rpc in queryset:
            # Add RPC number
            rpc_num = Paragraph('RPC No. ' + str(rpc.no_RPC), style=getSampleStyleSheet()['Heading2'])
            elements.append(rpc_num)

            # Add table
            data = [['NO.','TANGGAL', 'NAMA KEGIATAN', 'TOTAL AJUAN', 'NO CEK', 'NOMER BANK']]
            total = 0
            row_num = 1
            for cek in rpc.cek_set.all():
                for ajuan in cek.ajuan_terkait.all():  # Akses ajuan melalui ManyToMany
                    row = [
                        row_num,
                        cek.tanggal,
                        ajuan.nama_kegiatan if ajuan else 'Batal',
                        babel.numbers.format_currency(ajuan.total_ajuan, 'IDR', locale='id_ID') if ajuan else '',
                        cek.no_cek,
                        cek.nomer_bank_tertarik,

                    ]
                    data.append(row)
                    row_num += 1
                    if ajuan:
                        total += ajuan.total_ajuan

            data.append(['', 'JUMLAH', '',babel.numbers.format_currency(total, 'IDR', locale='id_ID')])

            # Create table
            table = Table(data, colWidths=[30,70, 300, 150, 80, 150])  # Atur lebar kolom di sini
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -2), 'LEFT'),
                ('ALIGN', (3, 1), (3, -2), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('WORDWRAP', (1, 0), (-1, -1), 100),
            ]))
            elements.append(table)

        space = Spacer(1, 10 * mm)
        elements.append(space)
        tanda_tangan = [['', '', '']]
        label_ttd = ['Disetujui', '                    ', 'Diperiksa', '                    ', 'Diajukan']
        tanda_tangan.append(label_ttd)
        space_ttd = ' '
        tanda_tangan.append(space_ttd)
        tanda_tangan.append(space_ttd)
        tanda_tangan.append(space_ttd)
        data_ttd = ['(H. Misbah Rosyadi)', '                    ', '(M. Dhiya Ulhaq, S.E)', '                    ',
                    '(Ita Aryani)']
        data_ttd_2 = ['Ketua Yayasan', '                    ', 'Direktur Keuangan', '                    ', 'Bendahara']
        tanda_tangan.append(data_ttd)
        tanda_tangan.append(data_ttd_2)
        table_ttd = Table(tanda_tangan)
        table_ttd.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table_ttd)

        doc.build(elements)
        return response

    export_as_pdf.short_description = "Export selected as PDF"

    from openpyxl import Workbook
    from django.http import HttpResponse

    def export_as_excel(modeladmin, request, queryset):
        # Membuat workbook dan worksheet baru
        wb = Workbook()
        ws = wb.active

        # Menambahkan header ke worksheet
        header = ['NO', 'TANGGAL', 'NAMA KEGIATAN', 'TOTAL AJUAN', 'NO CEK', 'NOMER BANK']
        ws.append(header)

        # Menambahkan data dari queryset ke worksheet
        for index, rpc in enumerate(queryset, start=1):
            for cek in rpc.cek_set.all():
                for ajuan in cek.ajuan_terkait.all():
                    row_data = row_data = [index, cek.tanggal, ajuan.nama_kegiatan if ajuan else 'Batal', ajuan.total_ajuan if ajuan else '', cek.no_cek, cek.nomer_bank_tertarik.nomer_bank_tertarik]
                    ws.append(row_data)

        # Mengatur nama file
        filename = "RPC.xlsx"

        # Membuat response Excel
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Menyimpan workbook ke response
        wb.save(response)

        return response

    export_as_excel.short_description = "Export selected as Excel"

class CekInLineBankTertarik(admin.TabularInline):
    model = Cek
    extra = 0
    fields = ['tanggal','no_cek', 'nomer_bank_tertarik', 'RPC', 'total_cek',]
    readonly_fields = fields
    can_delete = False

class DanaMasukInLineBankTertarik(admin.TabularInline):
    model = DanaMasuk
    extra = 0
    fields = ['waktu_masuk','uraian', 'bank_penerima', 'total_dana',]
    readonly_fields = fields
    can_delete = False

from django.http import FileResponse

class BankTertarikAdmin(admin.ModelAdmin):
    search_fields = ('nomer_bank_tertarik',)
    inlines = [CekInLineBankTertarik, DanaMasukInLineBankTertarik]
    actions = ['export_as_pdf', 'export_as_excel']
    list_display = ('nomer_bank_tertarik','no_cek_syc_bank_tertarik','total_cek_syc_bank_tertarik', 'dana_masuk_syc_bank_tertarik','total_dana_syc_bank_tertarik','total_selisih_syc_bank')
    list_per_page = 20

    def get_search_results(self, request, queryset, search_term):
        # Pemrosesan pencarian dalam model saat ini (RekapPencairanCek)
        results, use_distinct = super().get_search_results(request, queryset, search_term)

        # Pemrosesan pencarian dalam model Cek berdasarkan ForeignKey (RPC)
        cek_results = Cek.objects.filter(
            Q(no_cek__icontains=search_term)
        )
        cek_results = cek_results.values_list('nomer_bank_tertarik__id', flat=True)  # Ambil ID yang sesuai

        # Gabungkan hasil pencarian dari model saat ini dan model Cek
        results |= queryset.filter(id__in=cek_results)

        return results, use_distinct

    def no_cek_syc_bank_tertarik(self, bank_tertarik_obj):
        no_cek = []
        cek_related_to_bank_tertarik = Cek.objects.filter(nomer_bank_tertarik=bank_tertarik_obj)
        for cek in cek_related_to_bank_tertarik:
            no_cek.append(cek.no_cek)

        if no_cek:
            return ", ".join(no_cek)
        else:
            return '-'
        self.no_cek_self = no_cek

    no_cek_syc_bank_tertarik.short_description = 'No Cek'
    def total_cek_syc_bank_tertarik(self, bank_tertarik_obj):
        total_cek = []
        cek_related_to_bank_tertarik = Cek.objects.filter(nomer_bank_tertarik=bank_tertarik_obj)
        for cek in cek_related_to_bank_tertarik:
            total_cek.append(cek.total_cek)

        # Menghitung total
        total = sum(total_cek)
        self.total_cek_self = total

        # Mengonversi total ke format Rupiah
        total_rupiah = babel.numbers.format_currency(total, 'IDR', locale='id_ID')

        return total_rupiah if total else '-'

    total_cek_syc_bank_tertarik.short_description = 'Pengeluaran'

    def dana_masuk_syc_bank_tertarik(self, bank_pererima_obj):
        dana_masuk = []
        dana_masuk_to_bank_penerima = DanaMasuk.objects.filter(bank_penerima=bank_pererima_obj)
        for danamasuk in dana_masuk_to_bank_penerima:
            dana_masuk.append(danamasuk.uraian)

        if dana_masuk:
            return ", ".join(dana_masuk)
        else:
            return '-'
        self.dana_masuk_self = dana_masuk

    dana_masuk_syc_bank_tertarik.short_description = 'Dana Masuk'

    def total_dana_syc_bank_tertarik(self, bank_pererima_obj):
        dana_masuk = []
        dana_masuk_to_bank_penerima = DanaMasuk.objects.filter(bank_penerima=bank_pererima_obj)
        for danamasuk in dana_masuk_to_bank_penerima:
            dana_masuk.append(danamasuk.total_dana)

        # Menghitung total
        total = sum(dana_masuk)
        self.total_dana_self = total

        # Mengonversi total ke format Rupiah
        total_rupiah = babel.numbers.format_currency(total, 'IDR', locale='id_ID')

        return total_rupiah if total else '-'

    total_dana_syc_bank_tertarik.short_description = 'Pemasukkan'

    def total_selisih_syc_bank(self, bank_pererima_obj):
        # Menghitung total pemasukkan
        total_pemasukkan = \
        DanaMasuk.objects.filter(bank_penerima=bank_pererima_obj).aggregate(total_pemasukkan=Sum('total_dana'))[
            'total_pemasukkan'] or 0

        # Menghitung total pengeluaran
        total_pengeluaran = \
        Cek.objects.filter(nomer_bank_tertarik=bank_pererima_obj).aggregate(total_pengeluaran=Sum('total_cek'))[
            'total_pengeluaran'] or 0

        # Menghitung selisih (pengurangan)
        selisih = total_pemasukkan - total_pengeluaran

        # Mengonversi selisih ke format Rupiah
        selisih_rupiah = babel.numbers.format_currency(selisih, 'IDR', locale='id_ID')

        return selisih_rupiah if selisih else '-'
        self.selisih_self = selisih

    total_selisih_syc_bank.short_description = 'Selisih'

    def export_as_pdf(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Bank.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=10, rightMargin=5, topMargin=5,
                                bottomMargin=5)
        elements = []

        # Add title
        title_style = getSampleStyleSheet()['Heading1']
        title_style.alignment = TA_CENTER
        title = Paragraph('Bank', style=title_style)
        elements.append(title)

        for bank in queryset:
            # Add RPC number
            bank_num = Paragraph('Nama Bank ' + str(bank.nomer_bank_tertarik), style=getSampleStyleSheet()['Heading2'])
            elements.append(bank_num)

            # Add table
            data = [['NO', 'NO CEK', 'PENGELUARAN','DANA MASUK', 'PEMASUKKAN']]
            total_cek = 0
            total_dana = 0
            row_num = 1
            for cek in Cek.objects.filter(nomer_bank_tertarik=bank):
                row = [
                    row_num,
                    cek.no_cek,
                    babel.numbers.format_currency(cek.total_cek, 'IDR', locale='id_ID') if cek.total_cek else '',
                    '',
                    '',
                ]
                data.append(row)
                row_num += 1
                if cek.total_cek:
                    total_cek += cek.total_cek

            for dana in DanaMasuk.objects.filter(bank_penerima=bank):
                row = [
                    row_num,
                    '',
                    '',
                    dana.uraian,
                    babel.numbers.format_currency(dana.total_dana, 'IDR', locale='id_ID'),
                ]
                data.append(row)
                row_num += 1
                if dana.total_dana:
                    total_dana += dana.total_dana

            total = total_dana - total_cek  # Hitung total

            data.append(['','Total Cek', babel.numbers.format_currency(total_cek, 'IDR', locale='id_ID'),'Total Dana Masuk', babel.numbers.format_currency(total_dana, 'IDR', locale='id_ID')])
            data.append(['','SELISIH', '','', babel.numbers.format_currency(total, 'IDR', locale='id_ID')])

            # Create table
            table = Table(data, colWidths=[30, 150, 150, 150, 150,])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                # ('ALIGN', (2, 1), (2, -2), 'LEFT'),
                # ('ALIGN', (3, 1), (3, -2), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('WORDWRAP', (1, 0), (-1, -1), 100),
            ]))
            elements.append(table)

        doc.build(elements)
        return response

    export_as_pdf.short_description = "Export selected as PDF"

    def export_as_excel(self, request, queryset):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Bank.xlsx"'

        # Inisialisasi buku kerja Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bank"

        # Tambahkan judul
        title_font = Font(name='Helvetica-Bold', size=16)
        title = "Bank"
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        ws.merge_cells('A1:E1')

        # Header kolom
        headers = ["NO", "NO CEK", "PENGELUARAN", "DANA MASUK", "PEMASUKKAN"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws["{}2".format(col_letter)]
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        # Data dari queryset
        row_num = 3
        for bank in queryset:
            # Add table data
            total_cek = 0
            total_dana = 0

            for cek in Cek.objects.filter(nomer_bank_tertarik=bank):
                row = [
                    row_num - 2,
                    cek.no_cek,
                    cek.total_cek,
                    None,  # Kosongi kolom DANA MASUK
                    None,  # Kosongi kolom PEMASUKKAN
                ]
                ws.append(row)
                if cek.total_cek:
                    total_cek += cek.total_cek
                row_num += 1

            for dana in DanaMasuk.objects.filter(bank_penerima=bank):
                row = [
                    row_num - 2,
                    None,  # Kosongi kolom NO CEK
                    None,  # Kosongi kolom PENGELUARAN
                    dana.uraian,
                    dana.total_dana,
                ]
                ws.append(row)
                if dana.total_dana:
                    total_dana += dana.total_dana
                row_num += 1

            total = total_dana - total_cek
            ws.append(["", "Total Cek", total_cek, "Total Dana Masuk", total_dana])
            ws.append(["", "SELISIH", total])

        # Mengatur lebar kolom secara otomatis
        for column in ws.columns:
            max_length = 0
            column = get_column_letter(column[0].column)  # Get the column name
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Mengatur tinggi baris secara otomatis
        for row in ws.iter_rows():
            for cell in row:
                ws.row_dimensions[cell.row].height = 14

        # Simpan buku kerja Excel ke HttpResponse
        wb.save(response)

        return response

    export_as_excel.short_description = "Export selected as Excel"


class AjuanInLineUnitAjuan(admin.TabularInline):
    model = Ajuan
    extra = 0
    fields = ['nomor_pengajuan','nama_kegiatan',  'waktu_ajuan', 'penanggung_jawab', 'total_ajuan', 'RAPT']
    readonly_fields = fields
    can_delete = False
class UnitAjuanAdmin(admin.ModelAdmin):
    inlines = [AjuanInLineUnitAjuan]

admin.site.register(UnitAjuan, UnitAjuanAdmin)
admin.site.register(BankTertarik, BankTertarikAdmin)
admin.site.register(Ajuan, AjuanAdmin)
admin.site.register(DanaMasuk, DanaMasukAdmin)
admin.site.register(BuktiKasKeluar, BuktiKasKeluarAdmin)
admin.site.register(RekapPencairanCek, RekapPencairanCekAdmin)
admin.site.register(RekapAjuanPengambilanTabungan, RekapAjuanPengambilanTabunganAdmin)
admin.site.register(Cek, CekAdmin)
admin.site.site_header = 'Sistem Ajuan Bu Ita'
admin.site.site_title = 'Selamat Datang di Sistem Ajuan Bu Ita'
admin.site.index_title = 'Selamat datang di Dashboard Sistem Ajuan Bu Ita'